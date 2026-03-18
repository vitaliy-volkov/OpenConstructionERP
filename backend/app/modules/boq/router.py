"""BOQ API routes.

Endpoints:
    POST   /boqs/                    — Create a new BOQ
    GET    /boqs/?project_id=xxx     — List BOQs for a project
    GET    /boqs/{boq_id}            — Get BOQ with all positions
    PATCH  /boqs/{boq_id}            — Update BOQ metadata
    DELETE /boqs/{boq_id}            — Delete BOQ and all positions
    POST   /boqs/{boq_id}/positions  — Add a position to a BOQ
    PATCH  /positions/{position_id}  — Update a position
    DELETE /positions/{position_id}  — Delete a position
    POST   /boqs/{boq_id}/validate   — Validate a BOQ against configured rule sets
    GET    /boqs/{boq_id}/export/csv — Export BOQ positions as CSV
    GET    /boqs/{boq_id}/export/excel — Export BOQ positions as Excel (xlsx)
"""

import csv
import io
import uuid
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse

from app.dependencies import CurrentUserId, RequirePermission, SessionDep
from app.modules.boq.schemas import (
    BOQCreate,
    BOQResponse,
    BOQUpdate,
    BOQWithPositions,
    PositionCreate,
    PositionResponse,
    PositionUpdate,
)
from app.modules.boq.service import BOQService

router = APIRouter()


def _get_service(session: SessionDep) -> BOQService:
    return BOQService(session)


# ── BOQ CRUD ──────────────────────────────────────────────────────────────────


@router.post(
    "/boqs/",
    response_model=BOQResponse,
    status_code=201,
    dependencies=[Depends(RequirePermission("boq.create"))],
)
async def create_boq(
    data: BOQCreate,
    _user_id: CurrentUserId,
    service: BOQService = Depends(_get_service),
) -> BOQResponse:
    """Create a new Bill of Quantities."""
    boq = await service.create_boq(data)
    return BOQResponse.model_validate(boq)


@router.get(
    "/boqs/",
    response_model=list[BOQResponse],
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def list_boqs(
    project_id: uuid.UUID = Query(..., description="Filter BOQs by project"),
    offset: int = Query(default=0, ge=0),
    limit: int = Query(default=50, ge=1, le=100),
    service: BOQService = Depends(_get_service),
) -> list[BOQResponse]:
    """List all BOQs for a given project."""
    boqs, _ = await service.list_boqs_for_project(
        project_id, offset=offset, limit=limit
    )
    return [BOQResponse.model_validate(b) for b in boqs]


@router.get(
    "/boqs/{boq_id}",
    response_model=BOQWithPositions,
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def get_boq(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> BOQWithPositions:
    """Get a BOQ with all its positions and grand total."""
    return await service.get_boq_with_positions(boq_id)


@router.patch(
    "/boqs/{boq_id}",
    response_model=BOQResponse,
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def update_boq(
    boq_id: uuid.UUID,
    data: BOQUpdate,
    service: BOQService = Depends(_get_service),
) -> BOQResponse:
    """Update BOQ metadata (name, description, status)."""
    boq = await service.update_boq(boq_id, data)
    return BOQResponse.model_validate(boq)


@router.delete(
    "/boqs/{boq_id}",
    status_code=204,
    dependencies=[Depends(RequirePermission("boq.delete"))],
)
async def delete_boq(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> None:
    """Delete a BOQ and all its positions."""
    await service.delete_boq(boq_id)


# ── Position CRUD ─────────────────────────────────────────────────────────────


@router.post(
    "/boqs/{boq_id}/positions",
    response_model=PositionResponse,
    status_code=201,
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def add_position(
    boq_id: uuid.UUID,
    data: PositionCreate,
    service: BOQService = Depends(_get_service),
) -> PositionResponse:
    """Add a new position to a BOQ.

    The boq_id in the URL takes precedence over the body field.
    """
    # Override body boq_id with URL path parameter
    data.boq_id = boq_id
    position = await service.add_position(data)

    return PositionResponse(
        id=position.id,
        boq_id=position.boq_id,
        parent_id=position.parent_id,
        ordinal=position.ordinal,
        description=position.description,
        unit=position.unit,
        quantity=float(position.quantity),
        unit_rate=float(position.unit_rate),
        total=float(position.total),
        classification=position.classification,
        source=position.source,
        confidence=float(position.confidence) if position.confidence else None,
        cad_element_ids=position.cad_element_ids,
        validation_status=position.validation_status,
        metadata_=position.metadata_,
        sort_order=position.sort_order,
        created_at=position.created_at,
        updated_at=position.updated_at,
    )


@router.patch(
    "/positions/{position_id}",
    response_model=PositionResponse,
    dependencies=[Depends(RequirePermission("boq.update"))],
)
async def update_position(
    position_id: uuid.UUID,
    data: PositionUpdate,
    service: BOQService = Depends(_get_service),
) -> PositionResponse:
    """Update a BOQ position. Recalculates total if quantity or unit_rate changed."""
    position = await service.update_position(position_id, data)

    return PositionResponse(
        id=position.id,
        boq_id=position.boq_id,
        parent_id=position.parent_id,
        ordinal=position.ordinal,
        description=position.description,
        unit=position.unit,
        quantity=float(position.quantity),
        unit_rate=float(position.unit_rate),
        total=float(position.total),
        classification=position.classification,
        source=position.source,
        confidence=float(position.confidence) if position.confidence else None,
        cad_element_ids=position.cad_element_ids,
        validation_status=position.validation_status,
        metadata_=position.metadata_,
        sort_order=position.sort_order,
        created_at=position.created_at,
        updated_at=position.updated_at,
    )


@router.delete(
    "/positions/{position_id}",
    status_code=204,
    dependencies=[Depends(RequirePermission("boq.delete"))],
)
async def delete_position(
    position_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> None:
    """Delete a single position."""
    await service.delete_position(position_id)


# ── Validation ────────────────────────────────────────────────────────────────


def _build_rule_sets(
    project_rule_sets: list[str],
    classification_standard: str,
    region: str,
) -> list[str]:
    """Determine which validation rule sets to apply based on project config.

    Always includes the project's configured rule sets (default: ["boq_quality"]).
    Adds standard-specific rules based on classification_standard and region.

    Args:
        project_rule_sets: Explicit rule sets from project config.
        classification_standard: e.g. "din276", "nrm", "masterformat".
        region: e.g. "DACH", "UK", "US".

    Returns:
        Deduplicated list of rule set names.
    """
    rule_sets = list(project_rule_sets)

    # Add classification-standard-specific rules
    if classification_standard == "din276" and "din276" not in rule_sets:
        rule_sets.append("din276")
    if classification_standard == "nrm" and "nrm" not in rule_sets:
        rule_sets.append("nrm")
    if classification_standard == "masterformat" and "masterformat" not in rule_sets:
        rule_sets.append("masterformat")

    # Add region-specific rules
    if region.upper() == "DACH" and "gaeb" not in rule_sets:
        rule_sets.append("gaeb")
    if region.upper() == "UK" and "nrm" not in rule_sets:
        rule_sets.append("nrm")
    if region.upper() == "US" and "masterformat" not in rule_sets:
        rule_sets.append("masterformat")

    return rule_sets


@router.post(
    "/boqs/{boq_id}/validate",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def validate_boq(
    boq_id: uuid.UUID,
    session: SessionDep,
    service: BOQService = Depends(_get_service),
) -> dict[str, Any]:
    """Validate a BOQ against configured rule sets.

    Loads the BOQ with all positions, determines which validation rule sets
    to apply based on the project configuration, runs the validation engine,
    and returns a full validation report.
    """
    from app.core.validation.engine import validation_engine
    from app.modules.projects.repository import ProjectRepository

    # Load BOQ with positions
    boq_data = await service.get_boq_with_positions(boq_id)

    # Load project to get classification config
    project_repo = ProjectRepository(session)
    project = await project_repo.get_by_id(boq_data.project_id)
    if project is None:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Project not found for this BOQ",
        )

    # Convert positions to the format expected by validation rules
    positions_data = [
        {
            "id": str(pos.id),
            "ordinal": pos.ordinal,
            "description": pos.description,
            "quantity": pos.quantity,
            "unit_rate": pos.unit_rate,
            "classification": pos.classification,
        }
        for pos in boq_data.positions
    ]

    # Determine rule sets from project config
    rule_sets = _build_rule_sets(
        project_rule_sets=project.validation_rule_sets or ["boq_quality"],
        classification_standard=project.classification_standard or "din276",
        region=project.region or "DACH",
    )

    # Run validation
    report = await validation_engine.validate(
        data={"positions": positions_data},
        rule_sets=rule_sets,
        target_type="boq",
        target_id=str(boq_id),
        project_id=str(boq_data.project_id),
        region=project.region,
        standard=project.classification_standard,
    )

    # Build response: summary + full results
    summary = report.summary()
    summary["results"] = [
        {
            "rule_id": r.rule_id,
            "rule_name": r.rule_name,
            "severity": r.severity.value,
            "passed": r.passed,
            "message": r.message,
            "element_ref": r.element_ref,
            "suggestion": r.suggestion,
        }
        for r in report.results
    ]

    return summary


# ── Export (CSV / Excel) ──────────────────────────────────────────────────────


def _get_classification_code(classification: dict[str, Any]) -> str:
    """Extract the most relevant classification code for display.

    Checks din276, nrm, masterformat in order.
    """
    if not classification:
        return ""
    for key in ("din276", "nrm", "masterformat"):
        val = classification.get(key, "")
        if val:
            return str(val)
    # Fall back to the first available key
    for val in classification.values():
        if val:
            return str(val)
    return ""


@router.get(
    "/boqs/{boq_id}/export/csv",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def export_boq_csv(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> StreamingResponse:
    """Export BOQ positions as a CSV file."""
    boq_data = await service.get_boq_with_positions(boq_id)

    output = io.StringIO()
    writer = csv.writer(output)

    # Header row
    writer.writerow([
        "Pos.",
        "Description",
        "Unit",
        "Quantity",
        "Unit Rate",
        "Total",
        "Classification",
    ])

    # Position rows
    for pos in boq_data.positions:
        writer.writerow([
            pos.ordinal,
            pos.description,
            pos.unit,
            f"{pos.quantity:.2f}",
            f"{pos.unit_rate:.2f}",
            f"{pos.total:.2f}",
            _get_classification_code(pos.classification),
        ])

    # Grand total row
    writer.writerow([
        "",
        "Grand Total",
        "",
        "",
        "",
        f"{boq_data.grand_total:.2f}",
        "",
    ])

    content = output.getvalue()
    output.close()

    safe_name = boq_data.name.encode("ascii", errors="replace").decode("ascii").replace('"', "'")
    filename = f"{safe_name}.csv"

    return StreamingResponse(
        iter([content]),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )


@router.get(
    "/boqs/{boq_id}/export/excel",
    dependencies=[Depends(RequirePermission("boq.read"))],
)
async def export_boq_excel(
    boq_id: uuid.UUID,
    service: BOQService = Depends(_get_service),
) -> StreamingResponse:
    """Export BOQ positions as an Excel (xlsx) file with formatting."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, numbers
    from openpyxl.utils import get_column_letter

    boq_data = await service.get_boq_with_positions(boq_id)

    wb = Workbook()
    ws = wb.active
    ws.title = "BOQ"

    # ── Header row ────────────────────────────────────────────────────────
    headers = [
        "Pos.",
        "Description",
        "Unit",
        "Quantity",
        "Unit Rate",
        "Total",
        "Classification",
    ]
    bold_font = Font(bold=True)

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = bold_font

    # ── Position rows ─────────────────────────────────────────────────────
    number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1  # #,##0.00

    for row_idx, pos in enumerate(boq_data.positions, start=2):
        ws.cell(row=row_idx, column=1, value=pos.ordinal)
        ws.cell(row=row_idx, column=2, value=pos.description)
        ws.cell(row=row_idx, column=3, value=pos.unit)

        qty_cell = ws.cell(row=row_idx, column=4, value=pos.quantity)
        qty_cell.number_format = number_format

        rate_cell = ws.cell(row=row_idx, column=5, value=pos.unit_rate)
        rate_cell.number_format = number_format

        total_cell = ws.cell(row=row_idx, column=6, value=pos.total)
        total_cell.number_format = number_format

        ws.cell(
            row=row_idx,
            column=7,
            value=_get_classification_code(pos.classification),
        )

    # ── Grand total row ───────────────────────────────────────────────────
    total_row = len(boq_data.positions) + 2
    total_label = ws.cell(row=total_row, column=2, value="Grand Total")
    total_label.font = bold_font

    grand_total_cell = ws.cell(row=total_row, column=6, value=boq_data.grand_total)
    grand_total_cell.font = bold_font
    grand_total_cell.number_format = number_format

    # ── Auto-width columns ────────────────────────────────────────────────
    for col_idx in range(1, len(headers) + 1):
        max_length = len(str(headers[col_idx - 1]))
        for row in ws.iter_rows(
            min_row=2,
            max_row=total_row,
            min_col=col_idx,
            max_col=col_idx,
        ):
            for cell in row:
                val = cell.value
                if val is not None:
                    max_length = max(max_length, len(str(val)))
        # Add a small padding; cap at 60 to avoid excessively wide columns
        adjusted = min(max_length + 3, 60)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted

    # Align numeric columns to the right
    for row in ws.iter_rows(min_row=2, max_row=total_row, min_col=4, max_col=6):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")

    # ── Write to bytes buffer and return ──────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    safe_name = boq_data.name.encode("ascii", errors="replace").decode("ascii").replace('"', "'")
    filename = f"{safe_name}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
        },
    )
